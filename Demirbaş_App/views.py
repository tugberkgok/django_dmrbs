import os
import sqlite3
import openpyxl
import datetime
import locale
import operator
import re
from django.shortcuts import render, redirect, get_object_or_404
from xlsxwriter.workbook import Workbook
from .models import Worker, Device
from .forms import RegisterForm, LoginForm, DataForm, WorkerName
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages

locale.setlocale(locale.LC_ALL, '')
an = datetime.datetime.now()
tarih = datetime.datetime.strftime(an, '%c')

# Create your views here.
@login_required(login_url = 'login')
def Main(request):
    data = Worker.objects.all().order_by('person')
    keyword = request.GET.get("keyword")
    key = request.GET.get("key")
    if key:
        stok = Device.objects.filter(stok__contains=key)
        device = Device.objects.filter(device__contains=key)
        brand = Device.objects.filter(brand__contains=key)
        model = Device.objects.filter(model__contains=key)
        serial = Device.objects.filter(serial__contains=key)
        status = Device.objects.filter(status__contains=key)
        exp = Device.objects.filter(exp__contains=key)
        liste = []
        if stok:
            for idx in stok:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif device:
            for idx in device:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif brand:
            for idx in brand:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif model:
            for idx in model:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif serial:
            for idx in serial:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif status:
            for idx in status:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})
        elif exp:
            for idx in exp:
                datas = Worker.objects.filter(person=idx)
                liste.append(datas[0])
            return render(request, "Main.html", {"veri": liste})

        return redirect("main")

    if keyword:
        data = Worker.objects.filter(person__contains=keyword)
        return render(request, "Main.html", {"veri": data})

    return render(request, "Main.html", {"veri": data})

@login_required(login_url='login')
def update(request, id):
    datas = Device.objects.filter(person_id=id)
    person = Worker.objects.filter(id=id)
    superuser = Worker.objects.filter(superuser=1)
    devices = Device.objects.all()
    workers = Worker.objects.all()
    super = 1
    iz = 1
    username = request.user.username
    if person[0] in superuser:
        if username == superuser[0].person:
            sayman = 1
            return render(request, "update.html", {"datas": devices, "name": person[0], "workers": workers, "super": super, "sayman": sayman})
        else:
            return render(request, "update.html", {"datas": devices, "name": person[0], "workers": workers, "super": super})
    else:
        try:
            sayman = 1
            if datas[0] != " ":
                return render(request, "update.html", {"datas": datas, "name": person[0], "iz": iz, "sayman": sayman})
        except:
            return render(request, "update.html", {"name": person[0]})

@login_required(login_url = 'login')
def objectConfirm(request, id):
    device = Device.objects.filter(id=id)
    return render(request, "objectConfirm.html", {"devices": device})

@login_required(login_url = 'login')
def personConfirm(request, id):
    worker = Worker.objects.filter(id=id)
    return render(request, "personConfirm.html", {"workers": worker})

@login_required(login_url = 'login')
def delete(request, id):
    worker = get_object_or_404(Worker, id=id)
    sayman = Worker.objects.filter(superuser=1)
    conn = sqlite3.connect('db.sqlite3')
    query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(sayman[0])
    result1 = conn.cursor()
    result1.execute(query1)
    sid = result1.fetchone()
    conn = sqlite3.connect('db.sqlite3')
    cursor = conn.cursor()
    query3 = "UPDATE Demirbaş_App_device SET person_id_id = '{}' WHERE person_id_id = '{}'".format(sid[0], id)
    cursor.execute(query3)
    conn.commit()
    conn.close()
    worker.delete()
    return redirect("main")

def loginUser(request):
    form = LoginForm(request.POST or None)
    if form.is_valid():
        username = form.cleaned_data.get("username")
        password = form.cleaned_data.get("password")
        user = authenticate(username=username, password=password)
        if user is None:
            fail = 1
            return render(request, "login.html", {"form": form, "fail": fail})
        login(request, user)
        return redirect("main")
    return render(request, "login.html", {"form": form})

@login_required(login_url = 'login')
def logoutUser(request):
    logout(request)
    return redirect("/")

@login_required(login_url = 'login')
def addPerson(request):
    form = WorkerName(request.POST or None)
    if form.is_valid():
        conn = sqlite3.connect('db.sqlite3')
        query = "SELECT person FROM Demirbaş_App_worker WHERE person = '{}'".format(str(form["person"].value()))
        result = conn.cursor()
        result.execute(query)
        name = result.fetchone()
        conn.close()
        try:
            if str(form["person"].value()) != str(name[0]):
                form.save()
                messages.success(request, "Kişi Kaydedildi")
                return redirect("main")
            else:
                return render(request, "addPerson.html", {"form": form, "name": name})
        except:
            form.save()
            messages.success(request, "Kişi Kaydedildi")
            return redirect("main")
    else:
        return render(request, "addPerson.html", {"form": form})


@login_required(login_url = 'login')
def allData(request):
    datas = Device.objects.all()
    workers = Worker.objects.all()
    super = 1
    return render(request, "all.html", {"datas": datas, "workers": workers, "super": super})

@login_required(login_url = 'login')
def addData(request, id):
    person = Worker.objects.filter(id=id)
    form = DataForm(request.POST or None)
    if form.is_valid():
        conn = sqlite3.connect('db.sqlite3')

        query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(person[0])
        result = conn.cursor()
        result.execute(query1)
        pid = result.fetchone()

        query2 = "INSERT INTO Demirbaş_App_device (stok, device, number, brand, model, serial, status, exp, person_id_id) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')"\
                .format(str(form["stok"].value()), str(form["device"].value()), str(form["number"].value()), str(form["brand"].value()), str(form["model"].value()),
                        str(form["serial"].value()), str(form["status"].value()), str(form["exp"].value()), pid[0])
        c = conn.cursor()
        c.execute(query2)
        conn.commit()
        
        username1 = request.user.username
        query3 = "INSERT INTO Demirbaş_App_history (who, operation_type, stok, device, operation_date) VALUES ('{}', '{}', '{}', '{}', '{}')"\
                .format(username1,"Veriler Eklendi",str(form["stok"].value()), str(form["device"].value()),tarih)
        c = conn.cursor()        
        c.execute(query3)
        conn.commit()
        conn.close()

        return redirect("/update/{}".format(pid[0]))

    return render(request, "addData.html", {"form": form})


@login_required(login_url = 'login')
def objectEdit(request, id):
    datas = Device.objects.filter(id=id)
    conn = sqlite3.connect('db.sqlite3')
    query = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(datas[0])
    result = conn.cursor()
    result.execute(query)
    pid = result.fetchone()

    username1 = request.user.username 
    
    form = DataForm(request.POST or None, request.FILES or None, instance=get_object_or_404(Device, id=id))
    
    query3 = "INSERT INTO Demirbaş_App_history (who, operation_type, stok, device, operation_date) VALUES ('{}', '{}', '{}', '{}', '{}')"\
                .format(username1,"Veriler Düzenlendi",str(form["stok"].value()), str(form["device"].value()),tarih)
    c = conn.cursor()        
    c.execute(query3)
    conn.commit()
    conn.close()

    if form.is_valid():
        form.save()
        return redirect("/update/{}".format(pid[0]))

    return render(request, "objectEdit.html", {"form": form})

@login_required(login_url = 'login')
def objectDelete(request, id):
    object = Device.objects.filter(id=id)
    sayman = Worker.objects.filter(superuser=1)

    username1 = request.user.username
    print(object[0])

    conn = sqlite3.connect('db.sqlite3')
    cursor = conn.cursor()

    query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(sayman[0])
    result1 = conn.cursor()
    result1.execute(query1)
    sid = result1.fetchone()
    query2 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(object[0])
    result2 = conn.cursor()
    result2.execute(query2)
    pid = result2.fetchone()


    query4 = "SELECT stok, device FROM Demirbaş_App_device WHERE id = '{}'".format(id)
    Result4 = conn.cursor()
    Result4.execute(query4)
    data = Result4.fetchall()
    print(data)

    if str(object[0]) == str(sayman[0]):
        query3 = "INSERT INTO Demirbaş_App_history (who, operation_type, stok, device, operation_date) VALUES ('{}', '{}', '{}', '{}', '{}')"\
                .format(username1,"Veriler Silidi", data[0], data[1], tarih)
        c = conn.cursor()        
        c.execute(query3)
        conn.commit()
        conn.close()
        object.delete()
        return redirect("/update/{}".format(sid[0]))
    else:
        cursor = conn.cursor()
        query3 = "UPDATE Demirbaş_App_device SET person_id_id = '{}' WHERE id = '{}'".format(sid[0], id)
        cursor.execute(query3)
        conn.commit()

        query5 = "INSERT INTO Demirbaş_App_history (who, operation_type, stok, device, operation_date) VALUES ('{}', '{}', '{}', '{}', '{}')"\
                .format(username1,"Veriler Silindi ", data[0][0], data[0][1], tarih)
        c = conn.cursor()
        c.execute(query5)
        conn.commit()

        conn.close()
        return redirect("/update/{}".format(pid[0]))

@login_required(login_url = 'login')
def about(request):
    return render(request, "about.html")

@login_required(login_url = 'login')
def allExcelwrite(request):
    conn = sqlite3.connect('db.sqlite3')
    query = "SELECT stok, device, number, brand, model, serial, status, exp, iz, take_date, zim_date FROM Demirbaş_App_device"
    result = conn.cursor()
    result.execute(query)
    data = result.fetchall()
    workbook = Workbook(os.path.join(os.path.join(os.environ['USERPROFILE'])) + "\\Desktop\\Tüm Demirbaş.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.set_column(
        "D:D", 5

    )
    worksheet.set_column(
        "G:G",
        30
    )
    cell_format1 = workbook.add_format({'bold': True, 'italic': False})

    worksheet.write('D1', 'DEMİRBAŞ ENVANTER LİSTESİ', cell_format1)
    worksheet.write('A3', 'S.NO', cell_format1)
    worksheet.write('B3', 'STOK', cell_format1)
    worksheet.write('C3', 'CİHAZ', cell_format1)
    worksheet.write('D3', 'SAYI', cell_format1)
    worksheet.write('E3', 'MARKA', cell_format1)
    worksheet.write('F3', 'MODEL', cell_format1)
    worksheet.write('G3', 'SERİ NO', cell_format1)
    worksheet.write('H3', 'DURUMU', cell_format1)
    worksheet.write('I3', 'AÇIKLAMA', cell_format1)
    worksheet.write('J3', 'ESKİ KULLANICI / TARİH', cell_format1)
    worksheet.write('K3', 'FİYAT', cell_format1)
    worksheet.write('L3', 'ALIM TARİHİ', cell_format1)
    worksheet.write('M3', 'ZİMMET TARİHİ', cell_format1)

    row = 0
    cell = 0
    id = 1

    for row, veri in enumerate(data):
        for cell, value in enumerate(veri):
            if str(value) == "None":
                worksheet.write(row + 3, 0, id)
                worksheet.write(row + 3, cell, value)
            else:
                worksheet.write(row + 3, 0, id)
                worksheet.write(row + 3, cell + 1, value)
        id += 1

    workbook.close()
    conn.close()
    return redirect("allData")

@login_required(login_url = 'login')
def excelwrite(request, id):
    person = Worker.objects.filter(id=id)
    conn = sqlite3.connect('db.sqlite3')
    query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(person[0])
    result1 = conn.cursor()
    result1.execute(query1)
    pid = result1.fetchone()
    query2 = "SELECT stok, device, number, brand, model, serial, status, exp, iz, take_date, zim_date FROM Demirbaş_App_device WHERE person_id_id = '{}'".format(pid[0])
    result2 = conn.cursor()
    result2.execute(query2)
    data = result2.fetchall()
    workbook = Workbook(os.path.join(os.path.join(os.environ['USERPROFILE'])) + "\\Desktop\\{}.xlsx".format(person[0]))
    worksheet = workbook.add_worksheet()
    worksheet.set_column(
        "D:D", 5
    )
    worksheet.set_column(
        "G:G",
        30
    )
    cell_format1 = workbook.add_format({'bold': True, 'italic': False})

    worksheet.write('D1', 'DEMİRBAŞ ENVANTER LİSTESİ', cell_format1)
    worksheet.write('A3', 'S.NO', cell_format1)
    worksheet.write('B3', 'STOK', cell_format1)
    worksheet.write('C3', 'CİHAZ', cell_format1)
    worksheet.write('D3', 'SAYI', cell_format1)
    worksheet.write('E3', 'MARKA', cell_format1)
    worksheet.write('F3', 'MODEL', cell_format1)
    worksheet.write('G3', 'SERİ NO', cell_format1)
    worksheet.write('H3', 'DURUMU', cell_format1)
    worksheet.write('I3', 'AÇIKLAMA', cell_format1)
    worksheet.write('J3', 'ESKİ KULLANICI / TARİH', cell_format1)
    worksheet.write('K3', 'FİYAT', cell_format1)
    worksheet.write('L3', 'ALIM TARİHİ', cell_format1)
    worksheet.write('M3', 'ZİMMET TARİHİ', cell_format1)
    row = 0
    cell = 0
    id = 1
    for row, veri in enumerate(data):
        for cell, value in enumerate(veri):
            if str(value) == "None":
                worksheet.write(row + 3, 0, id)
                worksheet.write(row + 3, cell, value)
            else:
                worksheet.write(row + 3, 0, id)
                worksheet.write(row + 3, cell + 1, value)
        id += 1
    cell = 0
    row = row + 7
    worksheet.write(row, cell, "Yukarıda listelenen .......... kalem malzemeyi sağlam olarak teslim aldım.", cell_format1)
    worksheet.write(row + 2, cell, "Tesliim Alan: ", cell_format1)
    worksheet.write(row + 2, cell + 4, "Tesliim Eden: ", cell_format1)
    worksheet.write(row + 3, cell, "İmzası: ", cell_format1)
    worksheet.write(row + 3, cell + 4, "İmzası: ", cell_format1)
    worksheet.write(row + 4, cell, "Tarih:", cell_format1)
    worksheet.write(row + 4, cell + 4, "Tarih:", cell_format1)
    workbook.close()
    conn.close()
    return redirect("/update/{}".format(pid[0]))


def excelread(request, id):
    if request.method == 'POST':
        uploadedFile = request.FILES['file']
        workbook = openpyxl.load_workbook(uploadedFile)
        sheet = workbook["Sheet1"]
        max = -1
        excel_data = list()
        for row in sheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            max = max + 1

        num = 0
        conn = sqlite3.connect('db.sqlite3')
        for num in range(max):
            y = 0
            c = conn.cursor()
            query = "INSERT INTO Demirbaş_App_device (stok, device, number, brand, serial, status, exp, model, person_id_id, take_date, zim_date, price, iz) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')"\
                .format(excel_data[num+1][y+1], excel_data[num+1][y+2], excel_data[num+1][y+3], excel_data[num+1][y+4], excel_data[num+1][y+6], excel_data[num+1][y+7],
                        excel_data[num+1][y+8], excel_data[num+1][y+5], id, excel_data[num+1][y+11], excel_data[num+1][y+12], excel_data[num+1][y+10], excel_data[num+1][y+9])
            c.execute(query)
            conn.commit()
        conn.close()
        return redirect("/update/{}".format(id))

@login_required(login_url = 'login')
def dropdown(request, id, pid):
    person = Worker.objects.filter(id=pid)
    mainperson = Worker.objects.filter(superuser=1)

    conn = sqlite3.connect('db.sqlite3')
    query3 =  "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(mainperson[0])
    result3 = conn.cursor()
    result3.execute(query3)
    adam = result3.fetchone()
    query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(person[0])
    result1 = conn.cursor()
    result1.execute(query1)
    data = result1.fetchone()
    cursor = conn.cursor()
    query2 = "UPDATE Demirbaş_App_device SET person_id_id = '{}' WHERE id = {}".format(data[0], id)
    cursor.execute(query2)
    conn.commit()
    conn.close()
    return redirect("/update/{}".format(adam[0]))

@login_required(login_url='login')
def dropConfirm(request, id, pid):
    taker = Worker.objects.filter(id=pid)
    device = Device.objects.filter(id=id)
    mainperson = Worker.objects.filter(superuser=1)
    conn = sqlite3.connect('db.sqlite3')
    query3 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(mainperson[0])
    result3 = conn.cursor()
    result3.execute(query3)
    adam = result3.fetchone()
    return render(request, "dropConfirm.html", {"taker": taker[0], "device": device[0], "sayman": adam[0]})

@login_required(login_url = 'login')
def dropdownAll(request, id, pid):
    person = Worker.objects.filter(id=pid)
    mainperson = Worker.objects.filter(superuser=1)

    conn = sqlite3.connect('db.sqlite3')
    query3 =  "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(mainperson[0])
    result3 = conn.cursor()
    result3.execute(query3)
    adam = result3.fetchone()
    query1 = "SELECT id FROM Demirbaş_App_worker WHERE person = '{}'".format(person[0])
    result1 = conn.cursor()
    result1.execute(query1)
    data = result1.fetchone()
    cursor = conn.cursor()
    query2 = "UPDATE Demirbaş_App_device SET person_id_id = '{}' WHERE id = {}".format(data[0], id)
    cursor.execute(query2)
    conn.commit()
    conn.close()
    return redirect("allData")


@login_required(login_url='login')
def dropConfirmAll(request, id, pid):
    taker = Worker.objects.filter(id=pid)
    device = Device.objects.filter(id=id)
    mainperson = Worker.objects.filter(superuser=1)
    all = 1
    return render(request, "dropConfirm.html", {"taker": taker[0], "device": device[0], "all": all})


def register(request):
    form = RegisterForm(request.POST or None)
    if form.is_valid():
        username = form.cleaned_data.get("username")
        password = form.cleaned_data.get("password")
        key = form.cleaned_data.get("special_key")

        

        newUser = User(username=username)
        newUser.set_password(password)

        newUser.save()

        return redirect("login")

    context = {"form": form}    
    return render(request, "register.html", context)

def superregister(request):
    form = RegisterForm(request.POST or None)
    if form.is_valid():
        username = form.cleaned_data.get("username")
        password = form.cleaned_data.get("password")
        key = form.cleaned_data.get("special_key")
        person = str(username)
        newUser = User(username=username)
        newUser.set_password(password)
        newUser.save()

        conn = sqlite3.connect('db.sqlite3')
        c1 = conn.cursor()
        query1 = "UPDATE Demirbaş_App_worker SET superuser = {} WHERE superuser = {}".format(0, 1)
        c1.execute(query1)
        conn.commit()

        c2 = conn.cursor()
        query2 = "INSERT INTO Demirbaş_App_worker (person, superuser) VALUES ('{}', {})".format(person, 1)
        c2.execute(query2)
        conn.commit()

        return redirect("login")

    context = {"form": form}
    return render(request, "superregister.html", context)

@login_required(login_url = 'login')
def users(request):
    conn = sqlite3.connect('db.sqlite3')
    query = "SELECT id, last_login, username FROM auth_user"
    result = conn.cursor()
    result.execute(query)
    data = result.fetchall()
    datas = []
    for x in range(9):
        list = []
        for y in range(5):
            try:
                list.append(data[x][y])
            except:
                break
        datas.append(list)

    return render(request, "Users.html", {"datas": datas})

@login_required(login_url = 'login')
def userConfirm(request, id):
    conn = sqlite3.connect('db.sqlite3')
    query = "SELECT id FROM auth_user WHERE id = {}".format(id)
    result = conn.cursor()
    result.execute(query)
    user = result.fetchone()
    conn.commit()
    conn.close()
    return render(request, "userConfirm.html", {"user": user})

@login_required(login_url = 'login')
def userDelete(request, id):
    conn = sqlite3.connect('db.sqlite3')
    result = conn.cursor()
    query = "DELETE FROM auth_user WHERE id = {}".format(id)
    result = conn.execute(query)
    conn.commit()
    conn.close()
    return redirect("users")

def history(request):
    conn = sqlite3.connect('db.sqlite3')
    query = "SELECT who, operation_type, stok, device,operation_date FROM Demirbaş_App_history"
    result = conn.cursor()
    result.execute(query)
    data = result.fetchall()
    datas = []
    for x in range(100):
        list = []
        for y in range(5):
            try:
                list.append(data[x][y])
            except:
                break
        datas.append(list)

    return render(request, "history.html", {"datas": datas})
